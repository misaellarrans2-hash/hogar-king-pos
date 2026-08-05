from decimal import Decimal, InvalidOperation

from flask import Blueprint, flash, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from openpyxl import load_workbook

from app.decorators import roles_requeridos
from app.extensions import db
from app.models import (
    HistorialPrecio,
    MovimientoInventario,
    Producto,
    Rol,
    Sede,
    Stock,
)

inventario_bp = Blueprint("inventario", __name__, url_prefix="/inventario")

MOTIVOS_SALIDA = ["Producto dañado", "Producto vencido", "Ajuste de conteo", "Otro"]


def _sede_seleccionada():
    """Sede sobre la que se esta trabajando: la del usuario, o la elegida
    por querystring si es SUPERADMIN (que puede ver las 2 sedes)."""
    if current_user.ve_todas_las_sedes():
        sede_id = request.args.get("sede_id", type=int)
        if sede_id:
            return Sede.query.get_or_404(sede_id)
        return Sede.query.order_by(Sede.nombre).first()
    return current_user.sede


@inventario_bp.route("/")
@login_required
@roles_requeridos(Rol.SUPERADMIN, Rol.ADMINSEDE, Rol.BODEGUERO, Rol.ASESOR)
def listar():
    sede = _sede_seleccionada()
    stocks = []
    if sede:
        stocks = (
            Stock.query.filter_by(sede_id=sede.id)
            .join(Producto)
            .filter(Producto.activo.is_(True))
            .order_by(Producto.nombre)
            .all()
        )

    return render_template(
        "inventario/listar.html",
        sede=sede,
        stocks=stocks,
        sedes_visibles=current_user.sedes_visibles(),
    )


@inventario_bp.route("/productos/nuevo", methods=["GET", "POST"])
@login_required
@roles_requeridos(Rol.SUPERADMIN, Rol.ADMINSEDE)
def crear_producto():
    sede = _sede_seleccionada()

    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip()
        nombre = request.form.get("nombre", "").strip()
        precio_venta = _a_decimal(request.form.get("precio_venta"))
        costo = _a_decimal(request.form.get("costo")) if current_user.puede_ver_costos() else Decimal("0")
        stock_inicial = request.form.get("stock_inicial", type=int) or 0
        stock_minimo = request.form.get("stock_minimo", type=int) or 0

        if not codigo or not nombre or precio_venta is None:
            flash("Codigo, nombre y precio de venta son obligatorios.", "danger")
            return redirect(url_for("inventario.crear_producto"))

        if Producto.query.filter_by(codigo=codigo).first():
            flash(f"Ya existe un producto con el codigo {codigo}.", "danger")
            return redirect(url_for("inventario.crear_producto"))

        producto = Producto(codigo=codigo, nombre=nombre, precio_venta=precio_venta, costo=costo)
        db.session.add(producto)
        db.session.flush()

        db.session.add(
            Stock(producto_id=producto.id, sede_id=sede.id, cantidad=stock_inicial, stock_minimo=stock_minimo)
        )
        db.session.commit()

        flash(f'Producto "{nombre}" creado en {sede.nombre}.', "success")
        return redirect(url_for("inventario.listar", sede_id=sede.id))

    return render_template("inventario/producto_form.html", sede=sede, producto=None)


@inventario_bp.route("/productos/<int:producto_id>/editar", methods=["GET", "POST"])
@login_required
@roles_requeridos(Rol.SUPERADMIN, Rol.ADMINSEDE)
def editar_producto(producto_id):
    producto = Producto.query.get_or_404(producto_id)
    sede = _sede_seleccionada()

    if request.method == "POST":
        nombre = request.form.get("nombre", "").strip()
        precio_venta = _a_decimal(request.form.get("precio_venta"))

        if not nombre or precio_venta is None:
            flash("Nombre y precio de venta son obligatorios.", "danger")
            return redirect(url_for("inventario.editar_producto", producto_id=producto.id))

        precio_cambio = precio_venta != producto.precio_venta
        costo_cambio = False
        costo_nuevo = producto.costo

        if current_user.puede_ver_costos():
            costo_nuevo = _a_decimal(request.form.get("costo"))
            costo_cambio = costo_nuevo is not None and costo_nuevo != producto.costo

        if precio_cambio or costo_cambio:
            db.session.add(
                HistorialPrecio(
                    producto_id=producto.id,
                    usuario_id=current_user.id,
                    precio_anterior=producto.precio_venta,
                    precio_nuevo=precio_venta,
                    costo_anterior=producto.costo if costo_cambio else None,
                    costo_nuevo=costo_nuevo if costo_cambio else None,
                )
            )

        producto.nombre = nombre
        producto.precio_venta = precio_venta
        if costo_cambio:
            producto.costo = costo_nuevo

        db.session.commit()
        flash(f'Producto "{producto.nombre}" actualizado.', "success")
        return redirect(url_for("inventario.listar", sede_id=sede.id if sede else None))

    return render_template("inventario/producto_form.html", sede=sede, producto=producto)


@inventario_bp.route("/movimientos/entrada", methods=["GET", "POST"])
@login_required
@roles_requeridos(Rol.SUPERADMIN, Rol.ADMINSEDE, Rol.BODEGUERO)
def registrar_entrada():
    sede = _sede_seleccionada()

    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip()
        cantidad = request.form.get("cantidad", type=int)
        motivo = request.form.get("motivo", "Compra a proveedor").strip()

        producto = Producto.query.filter_by(codigo=codigo).first()

        if producto is None:
            flash(f"No existe ningun producto con el codigo {codigo}.", "danger")
            return redirect(url_for("inventario.registrar_entrada"))

        if not cantidad or cantidad <= 0:
            flash("La cantidad debe ser mayor a 0.", "danger")
            return redirect(url_for("inventario.registrar_entrada"))

        stock = producto.stock_en(sede.id)
        if stock is None:
            stock = Stock(producto_id=producto.id, sede_id=sede.id, cantidad=0, stock_minimo=0)
            db.session.add(stock)

        stock.cantidad += cantidad
        db.session.add(
            MovimientoInventario(
                producto_id=producto.id,
                sede_id=sede.id,
                usuario_id=current_user.id,
                tipo=MovimientoInventario.ENTRADA,
                cantidad=cantidad,
                motivo=motivo or "Compra a proveedor",
            )
        )
        db.session.commit()

        flash(f"Entrada registrada: +{cantidad} de {producto.nombre}. Stock actual: {stock.cantidad}.", "success")
        return redirect(url_for("inventario.registrar_entrada", sede_id=sede.id))

    return render_template("inventario/movimiento_form.html", sede=sede, tipo="entrada", motivos=None)


@inventario_bp.route("/movimientos/salida", methods=["GET", "POST"])
@login_required
@roles_requeridos(Rol.SUPERADMIN, Rol.ADMINSEDE, Rol.BODEGUERO)
def registrar_salida():
    sede = _sede_seleccionada()

    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip()
        cantidad = request.form.get("cantidad", type=int)
        motivo = request.form.get("motivo", "").strip()

        producto = Producto.query.filter_by(codigo=codigo).first()

        if producto is None:
            flash(f"No existe ningun producto con el codigo {codigo}.", "danger")
            return redirect(url_for("inventario.registrar_salida"))

        if not cantidad or cantidad <= 0:
            flash("La cantidad debe ser mayor a 0.", "danger")
            return redirect(url_for("inventario.registrar_salida"))

        stock = producto.stock_en(sede.id)
        if stock is None or stock.cantidad < cantidad:
            disponible = stock.cantidad if stock else 0
            flash(f"No hay suficiente stock de {producto.nombre} (disponible: {disponible}).", "danger")
            return redirect(url_for("inventario.registrar_salida"))

        stock.cantidad -= cantidad
        db.session.add(
            MovimientoInventario(
                producto_id=producto.id,
                sede_id=sede.id,
                usuario_id=current_user.id,
                tipo=MovimientoInventario.SALIDA,
                cantidad=cantidad,
                motivo=motivo or "Ajuste de conteo",
            )
        )
        db.session.commit()

        flash(f"Salida registrada: -{cantidad} de {producto.nombre}. Stock actual: {stock.cantidad}.", "success")
        return redirect(url_for("inventario.registrar_salida", sede_id=sede.id))

    return render_template("inventario/movimiento_form.html", sede=sede, tipo="salida", motivos=MOTIVOS_SALIDA)


@inventario_bp.route("/movimientos")
@login_required
@roles_requeridos(Rol.SUPERADMIN, Rol.ADMINSEDE, Rol.BODEGUERO)
def historial_movimientos():
    sede = _sede_seleccionada()
    movimientos = []
    if sede:
        movimientos = (
            MovimientoInventario.query.filter_by(sede_id=sede.id)
            .order_by(MovimientoInventario.creado_en.desc())
            .limit(200)
            .all()
        )
    return render_template("inventario/movimientos.html", sede=sede, movimientos=movimientos)


@inventario_bp.route("/historial-precios")
@login_required
@roles_requeridos(Rol.SUPERADMIN, Rol.ADMINSEDE)
def historial_precios():
    cambios = HistorialPrecio.query.order_by(HistorialPrecio.creado_en.desc()).limit(200).all()
    return render_template("inventario/historial_precios.html", cambios=cambios)


@inventario_bp.route("/importar", methods=["GET", "POST"])
@login_required
@roles_requeridos(Rol.SUPERADMIN, Rol.ADMINSEDE)
def importar():
    sede = _sede_seleccionada()

    if request.method == "POST":
        archivo = request.files.get("archivo")
        if not archivo or not archivo.filename.endswith((".xlsx", ".xlsm")):
            flash("Sube un archivo Excel (.xlsx) valido.", "danger")
            return redirect(url_for("inventario.importar"))

        try:
            libro = load_workbook(archivo, data_only=True)
            hoja = libro.active
        except Exception:
            flash("No se pudo leer el archivo. Verifica que sea un Excel valido.", "danger")
            return redirect(url_for("inventario.importar"))

        creados = 0
        actualizados = 0
        errores = 0

        filas = hoja.iter_rows(min_row=2, values_only=True)
        for fila in filas:
            if not fila or fila[0] is None:
                continue
            try:
                codigo = str(fila[0]).strip()
                nombre = str(fila[1]).strip()
                precio_venta = Decimal(str(fila[2]))
                costo = Decimal(str(fila[3])) if len(fila) > 3 and fila[3] is not None else Decimal("0")
                stock_inicial = int(fila[4]) if len(fila) > 4 and fila[4] is not None else 0
            except (IndexError, InvalidOperation, ValueError, TypeError):
                errores += 1
                continue

            producto = Producto.query.filter_by(codigo=codigo).first()
            if producto is None:
                producto = Producto(codigo=codigo, nombre=nombre, precio_venta=precio_venta, costo=costo)
                db.session.add(producto)
                db.session.flush()
                creados += 1
            else:
                producto.nombre = nombre
                actualizados += 1

            stock = producto.stock_en(sede.id)
            if stock is None:
                db.session.add(
                    Stock(producto_id=producto.id, sede_id=sede.id, cantidad=stock_inicial, stock_minimo=0)
                )
            else:
                stock.cantidad = stock_inicial

        db.session.commit()
        flash(
            f"Importacion completa en {sede.nombre}: {creados} productos nuevos, "
            f"{actualizados} actualizados, {errores} filas con error.",
            "success" if errores == 0 else "warning",
        )
        return redirect(url_for("inventario.listar", sede_id=sede.id))

    return render_template("inventario/importar.html", sede=sede)


def _a_decimal(valor):
    if valor is None or valor == "":
        return None
    try:
        return Decimal(valor)
    except InvalidOperation:
        return None
